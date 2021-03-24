from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import random

wb = load_workbook('1.xlsx')
ws = wb.get_sheet_by_name('2')
ws3 = wb.create_sheet('3')

matrix = [[] for _ in range(1, ws.max_column + 1)]

for col in range(1, ws.max_column):
    for col1 in range(col, ws.max_column + 1):
        s = 0
        for row in range(1, ws.max_row + 1):
            if ws.cell(column = col, row = row).value == ws.cell(column = col1, row = row).value:
                s += 1
                per = s / (ws.max_row) * 100
        print('Столбец {0} схож со столбцом {1} на {2}%'.format(get_column_letter(col), get_column_letter(col1), per))
        matrix[col-1].insert(col1-1, per)
        if col != col1:
            matrix[col1-1].insert(col-1, per)

for i in range(len(matrix)):
    for j in range(len(matrix[i])):
        if i != j:
            print(matrix[i][j], end = '\t')
            ws3.cell(column = j + 1, row = i + 1, value = matrix[i][j])
        else:
            matrix[i][j] = 0
            print('-', end = '\t')
            
    print()

ma = 0
for i in range(len(matrix)):
    for j in range(len(matrix[i])):
        if matrix[i][j] > ma:
            ma = matrix[i][j]
            n = i + 1
            m = j+ 1
print(get_column_letter(n), get_column_letter(m), ma)




wb.save('1.xlsx')