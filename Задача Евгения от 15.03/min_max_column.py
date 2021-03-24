from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import random

wb = load_workbook('1.xlsx')
ws = wb.active
ws2 = wb.create_sheet('2')

for col in range(1, ws.max_column + 1):
    column = ws[get_column_letter(col)]
    cells = [cell.value for cell in column]
    minimum = min(cells)
    maximum = max(cells)
    avg = sum(cells) / len(cells)
    for cell in cells:        
        ws2.cell(column = col, row = cells.index(cell) + 1, value = 'min' if cell < avg else 'max')

wb.save('1.xlsx')