from openpyxl import Workbook                                                           # Подключили модуль Workbook из модуля openpyxl
from openpyxl.utils import get_column_letter                                            # Подключили функцию get_column_letter из модуля openpyxl.utils

wb = Workbook()                                                                         # Создали экземпляр класса Workbook с именем wb

dest_filename = 'empty_book.xlsx'                                                       # Создали переменную для названия файла

ws1 = wb.active                                                                         # Получили активный лист в книге wb и положили в переменную ws1
ws1.title = "range names"                                                               # Задали листу ws1 заголовок "range names"

for row in range(40):                                                                   # Цикл для прохода по 40 строчкам
    ws1.append(range(600))                                                              # В каждую строку добавили 600 элементов

ws2 = wb.create_sheet(title="Pi")                                                       # Создали лист с названием Pi и положили в переменную ws2

ws2['F5'] = 3.14                                                                        # Записали в ячейку F5 листа ws2 значение 3.14

ws3 = wb.create_sheet(title="Data")                                                     # Создали лист с названием Data и положили в переменную ws3
for row in range(10, 20):                                                               # Цикл для прохода по строчкам 10 - 19
    for col in range(27, 54):                                                           # Цикл для прохода по столбцам 27 - 53
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))   # В ячейку по адресу (row, col) записывается название её столбца
print(ws3['AA10'].value)                                                                # Напечатали в терминал значние ячейки AA10 листа ws3
wb.save(filename = dest_filename)                                                       # Сохранили файл рядом со скриптом с названием, записанным в переменной dest_filename