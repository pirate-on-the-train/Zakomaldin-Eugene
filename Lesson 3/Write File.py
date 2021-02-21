from openpyxl import Workbook                   # Подключили модуль Workbook из модуля openpyxl
from openpyxl.utils import get_column_letter    # Подключили функцию get_column_letter из модуля openpyxl.utils

wb = Workbook()                                 # Создали экземпляр класса Workbook с именем wb

dest_filename = 'empty_book.xlsx'               # Создали переменную для названия файла

ws1 = wb.active                                 # Получили активный лист в книге wb и положили в переменную ws1
ws1.title = "range names"                       # Задали листу ws1 заголовок "range names"

ws1['A9'] = 5                                   # Записали в ячейку A9 листа ws1 значение 5

wb.save(dest_filename)                          # Сохранили файл рядом со скриптом с названием, записанным в переменной dest_filename