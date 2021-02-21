from openpyxl import load_workbook                  # Подключили функцию load_workbook из модуля openpyxl
wb = load_workbook(filename = 'empty_book.xlsx')    # В переменную wb загрузили файл empty_book.xlsx
sheet_ranges = wb['range names']                    # Получили диапазон значений в листе range names
print(sheet_ranges['D18'].value)                    # Напечатали в терминал значние ячейки D18