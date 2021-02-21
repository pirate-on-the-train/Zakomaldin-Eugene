from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'tt.xlsx')
sheet_ranges = wb['Отчет']
ma = 0
mi = 9999999999999999999
number = set({})

for row in range(2, 29):
    v = sheet_ranges.cell(column=3, row=row)
    number.add(v.value)

for row in range(2, 29):
    a = sheet_ranges[sheet_ranges.cell(column=8, row=row).coordinate].value
    b = sheet_ranges[sheet_ranges.cell(column=9, row=row).coordinate].value
    if b > ma:
        ma = b
    if a < mi:
        mi = a
print('Мин: ' + mi)
print()
print(sheet_ranges['A1'].value)